"""Excel upload and ingestion router."""
from __future__ import annotations

import io
import logging
from datetime import date, datetime
from typing import Any

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from sqlalchemy.orm import Session

from database import get_db
from middleware.role_guard import require_admin
from models.project import Project

logger = logging.getLogger(__name__)

router = APIRouter()

# Columns the importer looks for (case-sensitive, matching docs/excel-format.md)
REQUIRED_COLUMNS = {"Project Name", "Customer"}
OPTIONAL_COLUMNS = {
    "Description", "Start Date", "End Date", "Project Type",
    "Priority", "Progress", "Manager", "Practice", "Required Skills",
}
ALL_COLUMNS = REQUIRED_COLUMNS | OPTIONAL_COLUMNS

_STATUS_MAP = {
    "active": "Active",
    "in progress": "In Progress",
    "on hold": "On Hold",
    "completed": "Completed",
    "cancelled": "Cancelled",
}
_TYPE_MAP = {"paid": "Paid", "presales": "Presales", "internal": "Internal", "support": "Support"}
_PRIORITY_MAP = {"high": "High", "medium": "Medium", "low": "Low"}


def _parse_date(value: Any) -> date | None:
    if value is None:
        return None
    if isinstance(value, (date, datetime)):
        return value.date() if isinstance(value, datetime) else value
    try:
        return date.fromisoformat(str(value).strip())
    except ValueError:
        return None


def _parse_skills(value: Any) -> list[str]:
    if not value:
        return []
    return [s.strip() for s in str(value).split(",") if s.strip()]


def _normalize(raw: str | None, mapping: dict) -> str | None:
    if not raw:
        return None
    return mapping.get(str(raw).strip().lower())


@router.post("/upload")
def upload_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _role: str = Depends(require_admin),
) -> dict:
    """Upload a .xlsx file and upsert projects into the repository.

    Returns counts of inserted and updated rows.
    """
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=422, detail="Only .xlsx files are accepted")

    content = file.file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Could not parse Excel file: {exc}")

    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    # Validate required columns
    missing = REQUIRED_COLUMNS - set(headers)
    if missing:
        raise HTTPException(
            status_code=422,
            detail=f"Missing required columns: {sorted(missing)}. "
                   f"See docs/excel-format.md for the expected schema.",
        )

    col_index = {name: idx for idx, name in enumerate(headers) if name in ALL_COLUMNS}

    def cell(row, col_name):
        idx = col_index.get(col_name)
        return row[idx].value if idx is not None else None

    inserted = 0
    updated = 0

    for row in ws.iter_rows(min_row=2):
        name = cell(row, "Project Name")
        customer = cell(row, "Customer")
        if not name:
            continue  # skip blank rows

        name = str(name).strip()
        customer = str(customer).strip() if customer else ""

        existing = (
            db.query(Project)
            .filter(Project.name == name, Project.customer == customer)
            .first()
        )

        skills_raw = cell(row, "Required Skills")
        progress_raw = cell(row, "Progress")

        if existing:
            existing.description = str(cell(row, "Description") or existing.description or "")
            existing.start_date = _parse_date(cell(row, "Start Date")) or existing.start_date
            existing.end_date = _parse_date(cell(row, "End Date")) or existing.end_date
            existing.type = _normalize(cell(row, "Project Type"), _TYPE_MAP) or existing.type
            existing.priority = _normalize(cell(row, "Priority"), _PRIORITY_MAP) or existing.priority
            existing.progress_percent = int(progress_raw) if progress_raw is not None else existing.progress_percent
            existing.manager = str(cell(row, "Manager") or existing.manager or "")
            existing.practice = str(cell(row, "Practice") or existing.practice or "")
            if skills_raw:
                existing.required_skills = _parse_skills(skills_raw)
            existing.source = "excel"
            updated += 1
        else:
            project = Project(
                name=name,
                customer=customer,
                description=str(cell(row, "Description") or ""),
                start_date=_parse_date(cell(row, "Start Date")),
                end_date=_parse_date(cell(row, "End Date")),
                type=_normalize(cell(row, "Project Type"), _TYPE_MAP),
                priority=_normalize(cell(row, "Priority"), _PRIORITY_MAP),
                progress_percent=int(progress_raw) if progress_raw is not None else 0,
                manager=str(cell(row, "Manager") or ""),
                practice=str(cell(row, "Practice") or ""),
                required_skills=_parse_skills(skills_raw),
                source="excel",
                status="Active",
            )
            db.add(project)
            inserted += 1

    db.commit()
    logger.info("Excel import: %d inserted, %d updated", inserted, updated)
    return {"inserted": inserted, "updated": updated, "total": inserted + updated}


@router.get("/preview")
def preview_excel(
    limit: int = 50,
    db: Session = Depends(get_db),
) -> list[dict]:
    """Return the most recently imported Excel-sourced projects."""
    projects = (
        db.query(Project)
        .filter(Project.source == "excel")
        .order_by(Project.name)
        .limit(limit)
        .all()
    )
    return [_project_dict(p) for p in projects]


def _project_dict(p: Project) -> dict:
    return {
        "project_id": p.project_id,
        "name": p.name,
        "customer": p.customer,
        "status": p.status,
        "type": p.type,
        "priority": p.priority,
        "start_date": str(p.start_date) if p.start_date else None,
        "end_date": str(p.end_date) if p.end_date else None,
        "progress_percent": p.progress_percent,
        "description": p.description,
        "source": p.source,
        "required_skills": p.required_skills or [],
        "manager": p.manager,
        "practice": p.practice,
    }
